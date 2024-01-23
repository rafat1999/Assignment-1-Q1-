# Work from collecting the day name from google

from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import datetime

file_path = "G:\Selenium Test\selium test\Demo2\Excel.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet_names = workbook.sheetnames
print(sheet_names)
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.google.com/")
search_box = driver.find_element(By.NAME, "q")

day_name = datetime.datetime.now().strftime('%A')

for sheet_name in sheet_names:
    if sheet_name.lower() == day_name.lower():
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=4, max_col=5):
            for cell in row:
                cell.value = None

        for row in sheet.iter_rows(min_row= 2, values_only=True):
            value = row[2]

            if value is not None:
                search_box.clear()
                search_box.send_keys(value)
                time.sleep(1)
                suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li//span")

                long_suggestions = [suggestion.text for suggestion in suggestions if len(suggestion.text.split()) > 2]
                short_suggestions = [suggestion.text for suggestion in suggestions if len(suggestion.text.split()) <= 2]


                combine_1 = ", ".join(long_suggestions)
                combine_2 = ", ".join(short_suggestions)
                column_1 = 4
                column_2 = 5
                b_row = 2

                while sheet.cell(row=b_row, column=column_1).value is not None:
                    b_row += 1

                sheet.cell(row=b_row, column=column_1, value=combine_1)
                sheet.cell(row=b_row, column=column_2, value=combine_2)

                workbook.save(file_path)


driver.close()
print("Test Case Successfully Completed")
workbook.close()